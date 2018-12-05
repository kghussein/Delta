#! python3

# First we need to import the Modules

import os , openpyxl , csv , shutil , re , tkinter
import ListDelta_full_run_v5 as ld
from openpyxl.styles import *
from openpyxl.cell import *
import datetime
from tkinter import ttk
import pdb
from tkinter import *
from tkinter import filedialog

from openpyxl.styles import Color, Fill
from openpyxl.styles.borders import Border, Side

from openpyxl.styles import Alignment 

import tkinter.messagebox
from tkinter import messagebox

from PIL import Image,ImageTk



# color codes to be used Green = new, Red= removed, Yellow = modified, and empty = white

yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00',  fill_type='solid')
greenFill =  PatternFill(start_color='0000FF00', end_color='0000FF00',  fill_type='solid')
redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000',  fill_type='solid')
darkgreenFill = PatternFill(start_color='006400', end_color='006400',  fill_type='solid')
whiteFill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF',  fill_type='solid')

#define Borders
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))




#Defined functions:

def helpIndex():
    messagebox.showinfo("HELP","\nPlease contact CTO group for support\n")
    
    

# This is the testing function it is a dummy we can get it when we use new in the menu bar

def NewCallBack():
   messagebox.showinfo( "New", "Please choose the directories")
   start_Delta()
   root.quit
   return

# This Function for the about function in the help--> about menu bar

def AboutCall():
   messagebox.showinfo( "A.R.T Delta_software", "\n This Program has been creaated to help running the Delta changes between two different softwares. It is been devlopped to run the delta on Ericsson MME.\n Written By: Khalid Hussein.\n Softwareversion: 1.4.\n Owned By:CTO Group.")
   #messagebox.showinfo( "Delta_software", "\n Written By: Khalid Hussein.\n Softwareversion: 1.4.\n Owned By:CTO Group.")

# This Function for the dummy buttons in the menu bar for example save as at the menu bar
def donothing():
   filewin = Toplevel(root)
   button = Button(filewin, text="Do nothing button")
   button.pack()

# This Function for the new release directory name

def new_Release_directory():
    MMERELEASE  = filedialog.askdirectory()
    Release_entry.insert(0,MMERELEASE)
    print(MMERELEASE)

# This Function for the old release directory
def Old_directory_folder():
    old_path_entry.select_clear()
    old_path_entry.delete(0, END)
    oldDirectory = filedialog.askdirectory(initialdir = "c:/",title = "Select the old release folder")
    old_path_entry.insert(0,oldDirectory)

        

# This Function for the new release directory    
def New_directory_folder():
    new_path_entry.select_clear()
    new_path_entry.delete(0, END)
    newDirectory = filedialog.askdirectory(initialdir = "c:/",title = "Select the new release folder")
    new_path_entry.insert(0,newDirectory)
    

# This Function for where to save the Delta
def Saving_directory():
    Saving_entry.select_clear()
    Saving_entry.delete(0, END)
    SavingTheDelta = filedialog.askdirectory(initialdir = "c:/",title = "Select where you want to save")
    Saving_entry.insert(0,SavingTheDelta)
    print(newDirectory)

# This function compare's two files. explantion of the steps below:


# When we receive the Delta files from Ericsson it containes 10 or more files. the below code will differentaite between the old files name and the new files name and will match the first portion of the name
# also it will open an Excel sheet and name the first sheet as the front sheet with k =0. the rest of the tabs it will be created when we choose what portion of the name we will choose to name the tab
# Tab names can not be more than 35 charactors

def start_Delta():
    print ('Starting The Delta')

   

    try:
      oldDirectory =old_path_entry.get()
      MMEoldReleaseDir = os.listdir(r'%s' %oldDirectory)
      print(MMEoldReleaseDir)

    except:
    
      messagebox.showinfo( "Old Release Folder ", "Please choose a valid Old release folder")
      
      Old_directory_folder()
      


    try:
      newDirectory =new_path_entry.get()
      MMEnewReleaseDir = os.listdir(r'%s' %newDirectory)
      print(MMEnewReleaseDir)
    except:
    
      messagebox.showinfo( "New Release Folder ", "Please choose a valid New release folder")
      New_directory_folder()
      

    try:
      SavingTheDelta =Saving_entry.get()
      savein = os.listdir(r'%s' %SavingTheDelta)
      print(SavingTheDelta)

    except:
    
      messagebox.showinfo( "Old Release Folder ", "Please choose a valid Saving folder")
      Saving_directory()
      
    try:  
      MMERELEASE =Release_entry.get()
      if MMERELEASE=="":
          return messagebox.showinfo( "Saving As ", "Please Enter the Name you want to save as")
    except:
      messagebox.showinfo( "Saving As ", "Please Enter the Name you want to save as")
      
  
# opening the Excel sheet. and filling the first Tab. and set it as the Active Tab

    wb = openpyxl.Workbook()
    wb.create_sheet(index= 0 , title = 'Front Page')
    sheetTab = wb.get_sheet_by_name('Front Page')
    
    sheetTab.cell(row = 5, column = 4).value = str(MMERELEASE)
    sheetTab.cell(row = 6, column = 2).value = 'Author'
    sheetTab.cell(row = 6, column = 4).value = 'Khalid G Hussein'
    sheetTab.cell(row = 7, column = 4).value ='New'
    sheetTab.cell(row = 8, column = 4).value = 'Removed'
    sheetTab.cell(row = 9, column = 4).value = 'Modified'
    sheetTab.cell(row = 7 , column = 5).fill = greenFill
    sheetTab.cell(row = 8 , column = 5).fill = redFill
    sheetTab.cell(row = 9 , column = 5).fill = yellowFill


    sheetTab = wb.active

    
    sheetTab.column_dimensions['B'].width = 30
    sheetTab.column_dimensions['D'].width = 30
    sheetTab.column_dimensions['E'].width = 30
    
    k = 1

# here where we start the comparison between the names and compare betwwen the same naming of files.  

    for oldFile  in MMEoldReleaseDir:
  
        for newFile in MMEnewReleaseDir :
        
            if oldFile[1:6] == newFile[1:6]:
            
                   nameTAB = newFile.replace(newFile[-44: ] , 'Delta')
            
                  # pdb.set_trace()
                   
                   wb.create_sheet(index= k , title = nameTAB)
                   sheetTab = wb.get_sheet_by_name(nameTAB)
                   

                   alignment = Alignment(wrap_text=True, shrink_to_fit=True)
                 
# alignment properties
                 
                   sheetTab.column_dimensions['A'].width = 80
                   sheetTab.column_dimensions['C'].width = 80
                   
                   
                   sheetTab.cell(row = 1, column = 1).alignment = Alignment(wrapText=True, shrink_to_fit=True)
               
                   sheetTab.cell(row = 1, column = 1).value = oldFile[-44: ]
                   sheetTab.cell(row = 1, column = 3).value = newFile[-44: ]
                   sheetTab.cell(row = 1, column = 1).fill = darkgreenFill
                   sheetTab.cell(row = 1, column = 3).fill = darkgreenFill

                   sheetTab.cell(row = 2, column = 1).value = oldFile.replace(oldFile[-45: ] , ' ')
                   sheetTab.cell(row = 2, column = 3).value = newFile.replace(newFile[-45: ] , ' ')
                   sheetTab.cell(row = 2, column = 1).fill = darkgreenFill
                   sheetTab.cell(row = 2, column = 3).fill = darkgreenFill
                        
                   ld.compare(oldFile,newFile,sheetTab,oldDirectory,newDirectory,SavingTheDelta)

                   k = k+1
                   
            continue



    print(os.path.dirname(str(oldDirectory)))
                    

    os.chdir(r'%s' %SavingTheDelta)
    wb.save('%s.xlsx' %MMERELEASE )
    messagebox.showinfo( "Delta Run", "Successful Run .\n Done.")


# The Gui code below to get the directories and the release and the help menu 


root = Tk()
im=Image.open("C:/Users/kh012q/Desktop/Python/Logo-image.jpg")

photo=ImageTk.PhotoImage(im)  
cv = Canvas()  
root.title("Delta Files Directory")

master = ttk.Frame(root, padding="50 50 50 50")
master.grid(column=0, row=0, sticky=(N, W, E, S))
master.columnconfigure(0, weight=1)
master.rowconfigure(0, weight=1)

MMERELEASE = StringVar()
oldDirectory = StringVar()
newDirectory = StringVar()
SavingTheDelta = StringVar()


#The menu bar. not all of them are working now. some of them i set as Dummy.
menubar = Menu(root)
filemenu = Menu(menubar, tearoff=0)
filemenu.add_command(label="New", command=NewCallBack)
filemenu.add_command(label="Open", command=donothing)
filemenu.add_command(label="Save", command=donothing)
filemenu.add_command(label="Save as...", command=donothing)
filemenu.add_command(label="Close", command=root.destroy)

filemenu.add_separator()
menubar.add_cascade(label="File", menu=filemenu)

helpmenu = Menu(menubar, tearoff=0)
helpmenu.add_command(label="Help Index", command=helpIndex)
helpmenu.add_command(label="About...", command=AboutCall)
helpmenu.add_command(label="Exit", command=root.destroy)
menubar.add_cascade(label="Help", menu=helpmenu)



#Ending the menu bar code
# Start the browse button command line to enter the Directory names


Release_entry = ttk.Entry(master, width=90)
Release_entry.grid(column=1, row=5, sticky=(W, E))
ttk.Button(master, text="Browse", command= new_Release_directory)
ttk.Label(master, text="Enter the Release Name").grid(row=5, sticky=W)

    
old_path_entry = ttk.Entry(master, width=70)
old_path_entry.grid(column=1, row=2, sticky=(W, E))
ttk.Button(master, text="Browse", command= Old_directory_folder).grid(column=3, row=2, sticky=W)
ttk.Label(master, text="Enter The Old Release").grid(row=2, sticky=W)          

new_path_entry = ttk.Entry(master, width=70)
new_path_entry.grid(column=1, row=3, sticky=(W, E))
ttk.Button(master, text="Browse", command= New_directory_folder).grid(column=3, row=3, sticky=W)
ttk.Label(master, text="Enter The New Release").grid(row=3, sticky=W)


Saving_entry = ttk.Entry(master, width=70)
Saving_entry.grid(column=1, row=4, sticky=(W, E))
ttk.Button(master, text="Browse", command= Saving_directory).grid(column=3, row=4, sticky=W)
ttk.Label(master, text="Enter the Saving Directory").grid(row=4, sticky=W)


ttk.Button(master, text="Start The Delta", command= start_Delta).grid(column=1, row=8, sticky=(S))

print(os.path.dirname(str(oldDirectory)))

# Deleteing the Entry after we run the Program
old_path_entry.delete(0,END)
new_path_entry.delete(0,END)
Release_entry.delete(0,END)
root.bind('<Return>', master.destroy)
root.withdraw
root.config(menu=menubar)
root.mainloop()        

