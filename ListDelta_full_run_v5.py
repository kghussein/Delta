import re
import os, sys

# First we need to import the Modules

import os , openpyxl , csv , shutil , re , tkinter,colorama, colored
from openpyxl.styles import *
from openpyxl.cell import *
import datetime




# color codes to be used Green = new, Red= removed, Yellow = modified, and empty = white

yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00',  fill_type='solid')
greenFill =  PatternFill(start_color='0000FF00', end_color='0000FF00',  fill_type='solid')
redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000',  fill_type='solid')
darkgreenFill = PatternFill(start_color='006400', end_color='006400',  fill_type='solid')
whiteFill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF',  fill_type='solid')



#define Borders
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


#definning lists and parameters
list_old=[]
list_new=[]
list_all=[]
list_added=[]
list_old_only=[]
list_new_only=[]
list_old_only_copy=[]
list_new_only_copy=[]
list_common_before=[]
list_common=[]
longerlist=0
subcompare=[]
subcompare2=[]
subcompare3=[]
subcompare4=[]
string_diff=''
string_diff_new=''
string_diff_new_O=''



def compare(file1, file2,sheetTab,oldDirectory,newDirectory,SavingTheDelta):

  try:  

# Getting the Delta files from the nodes. Ericsson sends those files to us. it is their responsibility.     
    filedirectoryone = str(oldDirectory)+r'\%s' %file1
    filedirectorytwo = str(newDirectory)+r'\%s' %file2


# Doing the work on the new directory
    os.chdir(os.path.dirname(str(newDirectory)))
    



    # changing the files to a list

    with open(filedirectoryone) as o:
         list_old = [line.rstrip('\n') for line in o]

    with open(filedirectorytwo) as n:
         list_new = [line.rstrip('\n') for line in n]
 
    #Union set list, All the elements in both lists
    list_all= set(list_old) | set(list_new)
    list_all_copy= list(list_all)
    
    #Common elements from both set list
    list_common= set(list_old) & set(list_new)
   

    # Elements in one list set
    list_old_only= set(list_old).difference(set(list_new))
    list_old_only_copy= sorted(list(list_old_only))
    if list_old_only_copy==[]:
       list_old_only_copy.append('Old-Bogus')
    
    list_new_only= set(list_new).difference(set(list_old))
    list_new_only_copy= sorted(list(list_new_only))
    if list_new_only_copy==[]:
       list_new_only_copy.append('New-Bogus')
    # To get the index of the element in a set
    
    list_old_len=len(list_old)
    list_new_len=len(list_new)
    print('old: '+str(list_old_len))
    print('new: '+str(list_new_len))     

    # getting the longer list in order to iterate and cover all the elements
    global longerlist
    if list_old_len > list_new_len:

        longerlist= list_old_len
        #print('old: '+str(longerlist))
    else:
        longerlist= list_new_len
        #print('new: '+str(longerlist))
    #print(str(longerlist))
    

    # starting the logic of the delta

    # first to iterate between the two list to differentiate between elements
    global list_new_len
    global list_old_len
    global list_common
    global x
    x=3
    for i in range(longerlist):
        

        if list_new[i] in list_added:
           
           pass


        # if the first it in common that mean it is in both files so it equal.
        elif list_new[i] in list_common:


           sheetTab.cell(row = x, column = 1).value = list_new[i]
           sheetTab.cell(row = x, column = 3).value = list_new[i]
           
           sheetTab.cell(row = x, column = 1).alignment = Alignment(wrapText=True, shrink_to_fit=True)
           sheetTab.cell(row = x, column = 3).alignment = Alignment(wrapText=True, shrink_to_fit=True)
           sheetTab.cell(row = x, column = 3).border = thin_border
           sheetTab.cell(row = x, column = 1).border = thin_border
           
           sheetTab.cell(row = x, column = 4).alignment = Alignment(wrapText=True, shrink_to_fit=True)
           sheetTab.cell(row = x, column = 5).alignment = Alignment(wrapText=True, shrink_to_fit=True)

           list_added.append(list_new[i])
           x+=1
           pass

        # other wise we will break the line to paramters to check what is the exact paramteres percetage are equal
        
        else:
           
           subcompare=list_new[i].split()

           chosen_item=''
           modified_value=False
           new_param=True
           least_diff_new=0.3
           least_diff_old=0.3
           
           # we iterate through the old lines that it is not matched in common to verify if it is a modified line or removed line
           # we will loop in the old file with what we have from the new file to check what is the percentage
           for item in list_old_only_copy:
               subcompare2=item.split()

               diff_SC=set(subcompare).difference(set(subcompare2))
               diff_sc_t= list(diff_SC)
              
               diff_SC_O=set(subcompare2).difference(set(subcompare))
               diff_sc_O= list(diff_SC_O)
 
               
               result =  all(elem in diff_sc_t  for elem in subcompare)


               if len(diff_SC) / len(subcompare) < least_diff_new and len(diff_SC_O) / len(subcompare2) < least_diff_old:

                  chosen_item= item
                  modified_value=True
                  least_diff_new = len(diff_SC) / len(subcompare)
                  least_diff_old = len(diff_SC_O) / len(subcompare2)
                  string_diff_new_O= " ".join(diff_sc_O)
                  string_diff_new= " ".join(diff_sc_t)
                  new_param=False
                  continue

               elif modified_value==True and item ==list_old_only_copy[(len(list_old_only_copy))-1]:
                  list_added.append(list_new[i])
                  list_added.append(chosen_item)

                  sheetTab.cell(row = x, column = 1).value = chosen_item
                  sheetTab.cell(row = x, column = 3).value = list_new[i]
                  sheetTab.cell(row = x  , column = 1).fill = yellowFill
                  sheetTab.cell(row = x  , column = 3).fill = yellowFill
                  sheetTab.cell(row = x, column = 4).value = string_diff_new_O
                  sheetTab.cell(row = x, column = 5).value = string_diff_new
                  sheetTab.cell(row = x, column = 1).alignment = Alignment(wrapText=True, shrink_to_fit=True)
                  sheetTab.cell(row = x, column = 3).alignment = Alignment(wrapText=True, shrink_to_fit=True)
                  sheetTab.cell(row = x, column = 3).border = thin_border
                  sheetTab.cell(row = x, column = 1).border = thin_border
                  
                  sheetTab.cell(row = x, column = 4).alignment = Alignment(wrapText=True, shrink_to_fit=True)
                  sheetTab.cell(row = x, column = 5).alignment = Alignment(wrapText=True, shrink_to_fit=True)


                  list_new_only_copy.remove(list_new[i])
                  list_old_only_copy.remove(chosen_item)
                  x+=1
                  break

                     

               elif new_param==True and item ==list_old_only_copy[(len(list_old_only_copy))-1] :
                        
                
               
                  sheetTab.cell(row = x, column = 1).value = " New paramter---->"
                  sheetTab.cell(row = x, column = 3).value = list_new[i]
                  sheetTab.cell(row = x, column = 3).fill = greenFill
                  
                  sheetTab.cell(row = x, column = 3).alignment = Alignment(wrapText=True, shrink_to_fit=True)
                  sheetTab.cell(row = x, column = 1).alignment = Alignment(wrapText=True, shrink_to_fit=True)
                  sheetTab.cell(row = x, column = 3).border = thin_border
                  sheetTab.cell(row = x, column = 1).border = thin_border

                  sheetTab.cell(row = x, column = 4).alignment = Alignment(wrapText=True, shrink_to_fit=True)
                  sheetTab.cell(row = x, column = 5).alignment = Alignment(wrapText=True, shrink_to_fit=True)

                  list_new_only_copy.remove(list_new[i])
                  x+=1

               # for item that will loop first before it find a match on the modified or the removal
               else:
                  
                  continue


        if  i > list_old_len-1:
           continue
          
        elif list_old[i] in list_added:
          
           pass

        elif list_old[i] in list_common:
           

           sheetTab.cell(row = x, column = 1).value = list_old[i]
           sheetTab.cell(row = x, column = 3).value = list_old[i]
           
           sheetTab.cell(row = x, column = 1).alignment = Alignment(wrapText=True, shrink_to_fit=True)
           sheetTab.cell(row = x, column = 3).alignment = Alignment(wrapText=True, shrink_to_fit=True)
           sheetTab.cell(row = x, column = 3).border = thin_border
           sheetTab.cell(row = x, column = 1).border = thin_border



           sheetTab.cell(row = x, column = 4).alignment = Alignment(wrapText=True, shrink_to_fit=True)
           sheetTab.cell(row = x, column = 5).alignment = Alignment(wrapText=True, shrink_to_fit=True)
           

           list_added.append(list_old[i])
           #print('before it leave old common')
           x+=1
           pass

           

        else:
           
           # first we identify the line which it is not matching from the old file
           subcompare3=list_old[i].split() 
           
           chosen_item_new=''
           modified_value_old=False
           removable_param=True
           least_diff_new_t=0.3
           least_diff_old_t=0.4


           # Iterate through the new lines that it is not matched in common to verify if it is a modified line or removed line
           for item in list_new_only_copy:
               subcompare4=item.split()
               

               diff_SC_old=set(subcompare4).difference(set(subcompare3))
               diff_sc_old_t= list(diff_SC_old)

               diff_SC_old_O=set(subcompare3).difference(set(subcompare4))
               diff_sc_old_O= list(diff_SC_old_O)
               

               
               result_old =  all(elem in diff_sc_old_t  for elem in subcompare4)
               


               if len(diff_SC_old_O) / len(subcompare3) < least_diff_new_t and len(diff_sc_old_t) / len(subcompare4) < least_diff_old_t:

                  chosen_item_new= item
                  modified_value_old=True
                  least_diff_new_t = len(diff_SC_old_O) / len(subcompare3)
                  least_diff_old_t = len(diff_sc_old_t) / len(subcompare4)
                  string_diff= " ".join(diff_sc_old_t)
                  string_diff_old_O= " ".join(diff_sc_old_O)
                  removable_param=False
                  continue


               if modified_value_old==True and item == list_new_only_copy[(len(list_new_only_copy))-1]:

                  list_added.append(list_old[i])
                  list_added.append(chosen_item_new)
                  
                  sheetTab.cell(row = x, column = 1).value = list_old[i]
                  sheetTab.cell(row = x, column = 3).value = chosen_item_new
                  sheetTab.cell(row = x  , column = 1).fill = yellowFill
                  sheetTab.cell(row = x  , column = 3).fill = yellowFill
                  sheetTab.cell(row = x, column = 4).value = string_diff_old_O
                  sheetTab.cell(row = x, column = 5).value = string_diff
                  sheetTab.cell(row = x, column = 1).alignment = Alignment(wrapText=True, shrink_to_fit=True)
                  sheetTab.cell(row = x, column = 3).alignment = Alignment(wrapText=True, shrink_to_fit=True)
                  sheetTab.cell(row = x, column = 3).border = thin_border
                  sheetTab.cell(row = x, column = 1).border = thin_border

                  sheetTab.cell(row = x, column = 4).alignment = Alignment(wrapText=True, shrink_to_fit=True)
                  sheetTab.cell(row = x, column = 5).alignment = Alignment(wrapText=True, shrink_to_fit=True)                  
                  x+=1
                  


                  # we need to remove it so we do not loop into it again
                  list_old_only_copy.remove(list_old[i])
                  list_new_only_copy.remove(chosen_item_new)
                  break

           
               elif removable_param==True and item == list_new_only_copy[(len(list_new_only_copy))-1]:
               
                  #print(list_old[i]+str(x)+'<---removed parameter')
                  
                  
                  sheetTab.cell(row = x, column = 1).value = list_old[i]
                  sheetTab.cell(row = x, column = 3).value = "<---removed parameter"
                  sheetTab.cell(row = x, column = 1).fill = redFill
                  sheetTab.cell(row = x, column = 3).alignment = Alignment(wrapText=True, shrink_to_fit=True)
                  sheetTab.cell(row = x, column = 1).alignment = Alignment(wrapText=True, shrink_to_fit=True)
                  sheetTab.cell(row = x, column = 3).border = thin_border 
                  sheetTab.cell(row = x, column = 1).border = thin_border
                  
                  sheetTab.cell(row = x, column = 4).alignment = Alignment(wrapText=True, shrink_to_fit=True)
                  sheetTab.cell(row = x, column = 5).alignment = Alignment(wrapText=True, shrink_to_fit=True)


                  # we need to remove it so we do not loop into it again
                  
                  list_old_only_copy.remove(list_old[i])
                  x+=1
                  continue

               else:
                  continue

                

    # Below code is to make sure all the elements are been processed
    # by checking the number of the elements before the delta and the elements that has been processed after the Delta.
    # this is been done by creating another list and adding the elements that has been processed


    check_all2=  all(elem in list_all_copy  for elem in list_added)
    print(check_all2)
  
    check_all=  all(elem in list_added  for elem in list_all_copy)
    print(check_all)

    
  except Exception as e:
    print (e)


 
    

  


          
